import openpyxl

def main():
    wb = openpyxl.load_workbook('test.xlsx')
    
    # print(wb.sheetnames)
    
    # work_sheet
    # 1シート目が選択される
    ws = wb.active
    print(ws['A1'].value)

    print(ws.cell(row=1, column=1).value)

    # 列すべてを取り出す
    """
    for i in range(1,8):
        print(ws.cell(row=i, column=2).value)
    """

    # 行数が増えても対応できる書き方(listでcellをひとまとめに扱う)
    # columns->rowsにも変更可能
    # listはpython
    for cell in list(ws.columns)[0]:
        print(cell.value)



def save():
    # workbookオブジェクトを作成
    wb = openpyxl.Workbook()

    # アクティブなシートを取得
    sheet = wb.active
    sheet.title = 'test_sheet_1'

    # ファイルとして保存
    wb.save('test.xlsx')


def add():
    wb = openpyxl.Workbook()

    # シートを追加する
    wb.create_sheet()
    wb.create_sheet()

    # wb[sheetName]
    # wb.remove(sheet)
    wb.remove(wb['Sheet1'])
    print(wb.sheetnames)

def copy():
    # xlsxを開く
    wb = openpyxl.load_workbook('example.xlsx')

    ws = wb.active

    wb2 = openpyxl.load_workbook('test.xlsx')
    ws2 = wb2.active

    a = 1
    for cell in list(ws.columns)[1]:
        # cellの値をコピーしていく
        ws2.cell(row=a, column=1).value = cell.value

    wb2.save('test.xlsx')

def copycell():

    from copy import copy
    wb = openpyxl.Workbook()

    new_sheet = wb.create_sheet('copycellsheet')

    # copyされるシートをdefaultとする
    default_wb = openpyxl.load_workbook('example.xlsx')
    default_sheet = default_wb.active

    for row in default_sheet.rows:
        for cell in row:
            new_cell = new_sheet.cell(row=cell.row, column = cell.col_idx, value = cell.value)

            print(cell)

            if cell.has_style:
                print("this cell has a style")
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

            print(new_cell.value)

    wb.save('copiedlist.xlsx')
    
# 対象となるCellを抽出する
def search():
    wb = openpyxl.load_workbook('example.xlsx')

    # 抽出するシートを選択
    sheet = wb.get_sheet_by_name('list')

    # 行として抽出する
    for row in sheet.rows:
        # row[0]で一列目取得
        # print(row[1].value)
        if row[1].value == '〇':
            print(row[0].value)
        
        # for cell in row:
            # print(cell.value)
            
            # 〇を抽出
        

    # sheetnames = wb.get_sheet_names()
    # print(sheetnames)

def search2():
    wb = openpyxl.load_workbook('example.xlsx')

    # 抽出するシートを選択
    sheet = wb.get_sheet_by_name('Sheet1')

    # 行として抽出する
    for row in sheet.rows:
        # row[0]で一列目取得
        # print(row[2].value)
        print('debug')
        if row[2].value == 'SQL インジェクション':
            print('test')
            print(row[2])
        
def searchList():
    targetList = []
    targetList.append(1)
    targetList.append(2)
    print(targetList)

    if 1 in targetList:
        print('get')

if __name__ == '__main__':
    # main()
    # save()
    # add()
    # copy()
    # copycell()
    # search()
    # search2()
    # copyList()
    searchList()