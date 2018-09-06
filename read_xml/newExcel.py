import openpyxl
import glob
from xmlToJson import *

def sampleCreateExcel():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'test_sheet_1'
    wb.save('test.xlsx')
    glob.glob("*.xlsx")

# jsonからExcelのリストを作成
def jsonToListExcel():
    return

# dictionaryからリストを作成
def dictionaryToListExcel():
    xmlList = getAllXml()

    scenarioList = []
    for xmlFile in xmlList:
        scenario = createDictionary(xmlFile)
        scenarioList.append(scenario)
        
    # scenarioList:全シナリオの情報

    # excel生成
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'シナリオリスト'

    # 一番左上
    sheet['A1'] = '脆弱性/シナリオ'

    # 脆弱性のリストを作成
    vulsList = []

    # 脆弱性の種類と件数を埋める
    for index, scenario in enumerate(scenarioList):
        print(str(index + 1) + "番目")
        print(scenario["scenarioName"])

        # シナリオ名記述
        sheet.cell(row = 1, column = (index + 2)).value = scenario["scenarioName"]

        # 脆弱性名挿入
        # 重複したとき用に脆弱性のリストを作成
        # ここにあると二週目に初期化されてしまう
        # vulsList = []

        for vuls_index, vuls in enumerate(scenario["vuls"]):

            # 既出の脆弱性を排除
            if vuls["vulName"] in vulsList:
                # リストの何番目に格納されているか, 0始まり
                # 上書きする
                vuls_index = vulsList.index(vuls["vulName"])
            else:
                # 新規の脆弱性
                vulsList.append(vuls["vulName"])
                # 何番目の脆弱性か検出
                vuls_index = vulsList.index(vuls["vulName"])
                # 新しく項目を足す
                sheet.cell(row = vuls_index + 2, column = 1).value = vuls["vulName"]
                print(vuls["vulName"]+"を追加しました")
            
            sheet.cell(row = vuls_index + 2, column = index + 2).value = vuls["count"]
            # 数値データとして挿入
            sheet.cell(row = vuls_index + 2, column = index + 2).data_type = 'n'
        print(vulsList)
    wb.save('list.xlsx')

    return scenarioList

# レポートのエクセルを作成
def createReportExcel(scenarioList):
    
    print("-------createReportExcel----------")

    templateBook = openpyxl.load_workbook('template.xlsx')
    templateSheet = templateBook['template']

    dictVul = {}
    # テンプレートから行番号と脆弱性を抽出
    for row in templateSheet.rows:
        if row[2].value == None or row[2].value == '脆弱性分類':
            continue

        # これはやりすぎかも
        # vul = {"row": row[2].row, "vulName": row[2].value}
        
        # listにする必要ないな！
        # 半角をすべて抜く
        clearVal = (row[2].value).replace(' ', '').replace('(', '').replace('（', '').replace(')', '').replace('）', '')

        # 脆弱性の辞書を作成
        dictVul[clearVal] = row[2].row

        # templateList.append(tempVul)

    """
    # 星取表のシートを取得
    listBook = openpyxl.load_workbook('list.xlsx')
    listSheet = listBook['シナリオリスト']

    """
    """
    # search対象のリストを作成
    targetList = []

    for column in list(listSheet.columns)[0]:
        if column.value == '脆弱性/シナリオ':
            continue
        targetList.append(column.value)

    for cols in list(listSheet.columns):
        print(cols[2].value)
    return
    """
    print("-------------dictVul--------------")
    print(dictVul)
    print("-----------------------------------")
    
    # コピーするワークブックを作成
    from copy import copy
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    # 次に記入する行番号を貯める
    globalIndex = 1
    for index, scenario in enumerate(scenarioList):
        print(scenario["scenarioName"])

        # シナリオが変わったら2行空ける
        globalIndex += 2
        # シナリオ毎の処理
        # 検索対象のリストを作成
        new_sheet.cell(row = globalIndex, column = 1).value = scenario["scenarioName"]

        for vul_index, vul in enumerate(scenario["vuls"]):
            
            targetVul = vul["vulName"].replace(' ', '').replace('(', '').replace('（', '').replace(')', '').replace('）', '')
            # テンプレートに存在するかを確認
            if targetVul in dictVul:
                # 次の行に進める                
                globalIndex += 1
                print(targetVul + "はテンプレートの" + str(dictVul[targetVul]) + "行目に存在します")
                
                # 行数を取得
                targetRowNum = dictVul[targetVul]
                # 10列目までを取得
                targetRow = list(templateSheet.iter_rows(min_row = targetRowNum, max_row = targetRowNum, max_col = 10))
                print(targetRow)

                # listの中にさらにTupleがありその中にCellオブジェクトがある
                for rows in targetRow:
                    for cell in rows:
                        new_cell = new_sheet.cell(row = globalIndex, column = cell.col_idx, value=cell.value)
                        # 脆弱性検出数を挿入
                        count_cell = new_sheet.cell(row = globalIndex, column = 4, value = vul["count"])
                        count_cell.data_type = 'n'
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)
                
            else:
                print(vul["vulName"] + "はテンプレートには存在しませんでした")

    new_workbook.save('report.xlsx')



scenarioList = dictionaryToListExcel()
createReportExcel(scenarioList)