import openpyxl

def copyTemplate():
    templateBook = openpyxl.load_workbook('template.xlsx')

    # get_sheet_by_nameは今後消える
    templateSheet = templateBook['template']

    # 星取表のシートを取得
    listBook = openpyxl.load_workbook('list.xlsx')
    listSheet = listBook['list']

    # search対象のリストを作成
    targetList = []

    # 〇が付いたものをリストに格納
    for row in listSheet.rows:
        if row[1].value == '〇':
            targetList.append(row[0].value)

    # コピーするワークブックを作成
    from copy import copy
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    # 検索でヒットした回数
    hitCount = 0

    # targetListに該当するセルをtemplateシートからコピーする
    for row in templateSheet.rows:

        # listの中に含まれているか
        if row[2].value in targetList:

            # ヒット回数を上げる
            hitCount+=1

            # 含まれていたらコピーする
            for cell in row:
                new_cell = new_sheet.cell(row=hitCount,column=cell.col_idx, value=cell.value)

                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

    # 保存
    new_workbook.save('report.xlsx')


if __name__ == '__main__':

    copyTemplate()